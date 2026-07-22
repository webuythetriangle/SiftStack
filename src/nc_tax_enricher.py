"""Enrich NC notices with county tax delinquency data.

Unlike Knox County (tax_enricher.py), which has one REST API that covers
address lookup, parcel lookup, and delinquency in one place, each of the 5
NC target counties (config.ECOURTS_TARGET_COUNTIES) publishes delinquent
tax data through a completely different system. There is one lookup path
per county rather than a shared client.

Verified live 2026-07-22 (by downloading/querying each source directly —
see CLAUDE.md "NC Tax Delinquency Enrichment" for the full writeup):

  - Wake: daily bulk XLSX at services.wake.gov/collection_extracts/, with
    separate Street_Number/STREET_NAME columns — no address parsing needed.
  - Durham: property.spatialest.com's "Bill PWA" has an undocumented JSON
    endpoint (found by reading the SPA's main.js) that returns the full
    delinquent list in one call. HTTPS is required — the HTTP version of
    the same URL silently returns an empty body.
  - Guilford: delinquency lives in one ArcGIS FeatureServer keyed by REID
    with NO address field; address comes from a second ArcGIS FeatureServer
    (the county parcel layer), joined by REID.
  - Orange: an annual bulk XLSX (owner + parcel # + legal description, no
    address) joined by PIN against the county's ArcGIS parcel layer.
  - Mecklenburg: no automatable source found — see _mecklenburg_note_once().
    tax.mecknc.gov and taxbill.co.mecklenburg.nc.us both block automated
    fetches (robots.txt Content-Signal: ai-train=no, and the bill-search
    portal 403s outside a real browser session/WAF challenge). The GIS
    parcel layers (Polaris3G) carry no delinquency status. The only
    complete list is the annual "Delinquent Taxpayer List" the county is
    required to publish (NC GS 105-369) — available on request from the
    Office of Tax Administration, not as a scrapeable file.

A wrong first draft of the Guilford source (a same-schema ArcGIS layer
under a different org ID) turned out to be Virginia Beach, VA data, not
Guilford's — the URLs below were confirmed live against real Guilford/NC
addresses before being hardcoded here.
"""

import logging
import random
import re
import time
from collections import defaultdict
from pathlib import Path

import requests

import config
from notice_parser import NoticeData

logger = logging.getLogger(__name__)

REQUEST_TIMEOUT = 30
REQUEST_DELAY_MIN = 0.5
REQUEST_DELAY_MAX = 1.2
_HEADERS = {
    "User-Agent": "SiftStackBot/1.0 (real estate lead research; "
    "contact vishal@webuythetriangle.com)"
}

NC_TAX_COUNTIES = {"wake", "durham", "orange", "guilford", "mecklenburg"}


# ── Address Matching ────────────────────────────────────────────────────

_STREET_SUFFIXES = {
    "ST", "STREET", "AVE", "AVENUE", "RD", "ROAD", "DR", "DRIVE", "LN",
    "LANE", "CT", "COURT", "BLVD", "BOULEVARD", "WAY", "PL", "PLACE",
    "CIR", "CIRCLE", "PIKE", "TRL", "TRAIL", "LOOP", "RUN", "TER",
    "TERRACE", "TERR", "PKWY", "PARKWAY", "HWY", "HIGHWAY", "XING",
    "CROSSING", "SQ", "SQUARE", "PT", "POINT", "COVE", "CV", "PASS",
    "WALK", "ROW", "BND", "BEND", "GLN", "GLEN", "RDG", "RIDGE", "VW",
    "VIEW", "MNR", "MANOR", "EXT", "EXTENSION", "HTS", "HEIGHTS",
}
_DIRECTIONALS = {"N", "S", "E", "W", "NE", "NW", "SE", "SW",
                  "NORTH", "SOUTH", "EAST", "WEST"}


def _address_key(address: str) -> tuple[str, str] | None:
    """(house_number, first_street_word) join key for matching notice
    addresses against county tax rolls.

    Loose by design — county files carry the official situs address while
    scraped notice addresses are free-text — so house number + first
    significant street word is used as a good-enough join key rather than
    an exact string match. Also guarantees the output is bare alphanumerics,
    safe to interpolate into an ArcGIS `where` clause without SQL escaping.
    """
    if not address:
        return None
    tokens = re.findall(r"[A-Za-z0-9]+", address.upper())
    if not tokens or not tokens[0].isdigit():
        return None
    house_number = tokens[0]
    for tok in tokens[1:]:
        if tok in _DIRECTIONALS or tok in _STREET_SUFFIXES or tok.isdigit():
            continue
        return (house_number, tok)
    return None


# ── Wake County ──────────────────────────────────────────────────────────

WAKE_BASE_URL = "https://services.wake.gov/collection_extracts"


def _wake_download() -> Path | None:
    """Download (or reuse a cached copy of) Wake's daily delinquent-tax XLSX.

    Refreshed daily under a date-stamped filename — try today, then walk
    back a few days in case today's file hasn't posted yet.
    """
    from datetime import datetime, timedelta

    for days_back in range(6):
        day = datetime.now() - timedelta(days=days_back)
        fname = day.strftime("REAL_ESTATE_delq853_%m%d%Y.xlsx")
        cache_path = config.NC_TAX_CACHE_DIR / fname
        if cache_path.exists():
            return cache_path
        try:
            resp = requests.get(
                f"{WAKE_BASE_URL}/{fname}", timeout=REQUEST_TIMEOUT, headers=_HEADERS
            )
            if resp.status_code == 200 and resp.content:
                cache_path.write_bytes(resp.content)
                logger.info("Wake: downloaded %s", fname)
                return cache_path
        except requests.RequestException as e:
            logger.debug("Wake: bulk download failed for %s: %s", fname, e)
    return None


def _wake_build_index() -> dict[tuple[str, str], dict]:
    """Parse Wake's bulk file into an address-keyed delinquency index.

    One row per (account, tax year) bill — grouped by ACCOUNT_NUM for a
    total amount due and a count of delinquent years per property.
    """
    path = _wake_download()
    if not path:
        return {}

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col = {name: i for i, name in enumerate(header)}

    required = {
        "ACCOUNT_NUM", "TAX_YEAR", "Street_Number", "STREET_NAME",
        "Primary_Owner", "TOTAL_DUE",
    }
    if not required.issubset(col):
        logger.warning(
            "Wake: unexpected column layout (missing %s) — skipping",
            required - set(col),
        )
        return {}

    by_account: dict[str, dict] = {}
    for row in rows:
        account = row[col["ACCOUNT_NUM"]]
        if not account:
            continue
        due = row[col["TOTAL_DUE"]] or 0
        if due <= 0:
            continue
        entry = by_account.setdefault(account, {
            "amount": 0.0,
            "years": set(),
            "owner": row[col["Primary_Owner"]] or "",
            "house_number": str(row[col["Street_Number"]] or "").strip(),
            "street": str(row[col["STREET_NAME"]] or "").strip().upper(),
        })
        entry["amount"] += float(due)
        entry["years"].add(row[col["TAX_YEAR"]])

    index: dict[tuple[str, str], dict] = {}
    for entry in by_account.values():
        house_number = entry["house_number"]
        street_tokens = entry["street"].split()
        if not house_number or not street_tokens:
            continue
        index[(house_number, street_tokens[0])] = {
            "delinquent_amount": round(entry["amount"], 2),
            "delinquent_years": len(entry["years"]),
            "owner": entry["owner"],
        }
    logger.info("Wake: indexed %d delinquent properties", len(index))
    return index


# ── Durham County ────────────────────────────────────────────────────────

DURHAM_API_URL = "https://property.spatialest.com/nc/durham-tax/data/getData.php"
_DURHAM_ADDR_RE = re.compile(r"^(\d+)\s+(.+?)\s+DURHAM\s+NC\s+\d{5}", re.IGNORECASE)


def _durham_extract_bill_year(bill: str) -> str | None:
    """Bill numbers look like '0000118976-2025-2025-0000-00' — 2nd segment is the tax year."""
    parts = (bill or "").split("-")
    if len(parts) >= 2 and parts[1].isdigit() and len(parts[1]) == 4:
        return parts[1]
    return None


def _durham_build_index() -> dict[tuple[str, str], dict]:
    """Fetch Durham's full delinquent list in one call and index by address.

    No auth/session needed — a stateless POST returns the whole list
    (~3000 bills as of 2026-07). Must be HTTPS; the HTTP URL returns an
    empty body with the same 200 status.
    """
    try:
        resp = requests.post(
            DURHAM_API_URL,
            data={"qtype": "delinquint_list", "search": ""},
            timeout=REQUEST_TIMEOUT,
            headers=_HEADERS,
        )
        resp.raise_for_status()
        results = resp.json().get("results", [])
    except (requests.RequestException, ValueError) as e:
        logger.warning("Durham: delinquent list fetch failed: %s", e)
        return {}

    by_parcel: dict[str, dict] = {}
    for row in results:
        parcel = row.get("Parcel") or row.get("PARID")
        if not parcel:
            continue
        try:
            amount = float(row.get("BalanceNum") or row.get("TotAdvBalanceDue") or 0)
        except (TypeError, ValueError):
            amount = 0.0
        addr_raw = (row.get("AssetDescription") or "").strip()
        m = _DURHAM_ADDR_RE.match(addr_raw)
        entry = by_parcel.setdefault(parcel, {
            "amount": 0.0,
            "years": set(),
            "bills": 0,
            "owner": row.get("NewOwner") or "",
            "house_number": m.group(1) if m else "",
            "street": m.group(2).strip().upper() if m else "",
        })
        entry["amount"] += amount
        entry["bills"] += 1
        year = _durham_extract_bill_year(row.get("Bill", ""))
        if year:
            entry["years"].add(year)

    index: dict[tuple[str, str], dict] = {}
    for entry in by_parcel.values():
        street_tokens = entry["street"].split()
        if not entry["house_number"] or not street_tokens:
            continue
        index[(entry["house_number"], street_tokens[0])] = {
            "delinquent_amount": round(entry["amount"], 2),
            # Distinct tax years when the bill number parses cleanly,
            # otherwise fall back to a bill count as a rough proxy.
            "delinquent_years": len(entry["years"]) or entry["bills"],
            "owner": entry["owner"],
        }
    logger.info("Durham: indexed %d delinquent properties", len(index))
    return index


# ── Guilford County ──────────────────────────────────────────────────────
# Delinquency table has no address field — resolved via a second
# ArcGIS layer (the county parcel/cadastral layer) joined on REID.

GUILFORD_DELQ_URL = (
    "https://services5.arcgis.com/RR1v7NWFfwk98pUn/arcgis/rest/services/"
    "Tax_Delinquent_Report_/FeatureServer/0/query"
)
GUILFORD_PARCEL_URL = (
    "https://gcgis.guilfordcountync.gov/arcgis/rest/services/"
    "GC_Cadastral_Current/GC_Parcels/FeatureServer/0/query"
)


def _guilford_fetch_all_delinquent() -> list[dict]:
    features: list[dict] = []
    offset = 0
    page_size = 2000
    while True:
        params = {
            "where": "1=1",
            "outFields": "PARCEL_NUM,OWNER_NAME,TAX_YEAR,TOTAL_DUE_AMOUNT,BILL_STATUS",
            "f": "json",
            "resultOffset": offset,
            "resultRecordCount": page_size,
        }
        try:
            resp = requests.get(
                GUILFORD_DELQ_URL, params=params, timeout=REQUEST_TIMEOUT, headers=_HEADERS
            )
            resp.raise_for_status()
            data = resp.json()
        except (requests.RequestException, ValueError) as e:
            logger.warning("Guilford: delinquent table fetch failed at offset %d: %s", offset, e)
            break
        batch = data.get("features", [])
        if not batch:
            break
        features.extend(f["attributes"] for f in batch)
        if len(batch) < page_size:
            break
        offset += page_size
        time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))
    return features


def _guilford_build_delinquent_by_reid() -> dict[str, dict]:
    rows = _guilford_fetch_all_delinquent()
    by_reid: dict[str, dict] = {}
    for row in rows:
        reid = row.get("PARCEL_NUM")
        if not reid:
            continue
        status = (row.get("BILL_STATUS") or "").upper()
        if status and status != "UNPAID":
            continue
        entry = by_reid.setdefault(
            reid, {"amount": 0.0, "years": set(), "owner": row.get("OWNER_NAME") or ""}
        )
        entry["amount"] += float(row.get("TOTAL_DUE_AMOUNT") or 0)
        if row.get("TAX_YEAR"):
            entry["years"].add(row["TAX_YEAR"])
    logger.info("Guilford: indexed %d delinquent parcels", len(by_reid))
    return by_reid


def _guilford_lookup_reids_by_address(address: str) -> list[str]:
    """Return candidate REIDs for an address — sometimes more than one

    (e.g. adjoining/resubdivided parcels sharing a street address), so the
    caller checks each against the delinquent index rather than assuming
    the first result is the right one.
    """
    key = _address_key(address)
    if not key:
        return []
    house_number, street_word = key
    where = f"PHYADDR_STR_NUM='{house_number}' AND PHYADDR_STR LIKE '{street_word}%'"
    params = {"where": where, "outFields": "REID", "f": "json", "resultRecordCount": 5}
    try:
        resp = requests.get(
            GUILFORD_PARCEL_URL, params=params, timeout=REQUEST_TIMEOUT, headers=_HEADERS
        )
        resp.raise_for_status()
        features = resp.json().get("features", [])
        return [f["attributes"].get("REID") for f in features if f["attributes"].get("REID")]
    except (requests.RequestException, ValueError) as e:
        logger.debug("Guilford: parcel address lookup failed for '%s': %s", address, e)
        return []


def _apply_guilford(notices: list[NoticeData]) -> None:
    by_reid = _guilford_build_delinquent_by_reid()
    if not by_reid:
        logger.info("Guilford: no delinquency data available this run")
        return
    enriched = 0
    for n in notices:
        time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))
        reid = next(
            (r for r in _guilford_lookup_reids_by_address(n.address) if r in by_reid), None
        )
        if not reid:
            continue
        result = by_reid[reid]
        n.tax_delinquent_amount = str(round(result["amount"], 2))
        n.tax_delinquent_years = str(len(result["years"]) or 1)
        if result.get("owner") and not n.tax_owner_name:
            n.tax_owner_name = result["owner"]
        n.parcel_id = n.parcel_id or reid
        enriched += 1
    logger.info("Guilford: tax-delinquent %d/%d", enriched, len(notices))


# ── Orange County ────────────────────────────────────────────────────────
# Annual bulk file has no address either — resolved via the county's
# ArcGIS parcel layer, joined on PIN.

ORANGE_BULK_URL = "https://www.orangecountync.gov/DocumentCenter/View/27374"
ORANGE_PARCEL_URL = "https://gis.orangecountync.gov/arcgis/rest/services/WebParcelService/MapServer/0/query"
ORANGE_CACHE_MAX_AGE_DAYS = 30  # published once/year (~March) — a month-old cache is fine


def _orange_download() -> Path | None:
    cache_path = config.NC_TAX_CACHE_DIR / "orange_delinquent.xlsx"
    if cache_path.exists():
        age_days = (time.time() - cache_path.stat().st_mtime) / 86400
        if age_days < ORANGE_CACHE_MAX_AGE_DAYS:
            return cache_path
    try:
        resp = requests.get(ORANGE_BULK_URL, timeout=REQUEST_TIMEOUT, headers=_HEADERS)
        resp.raise_for_status()
        cache_path.write_bytes(resp.content)
        logger.info("Orange: downloaded delinquent tax list")
        return cache_path
    except requests.RequestException as e:
        logger.warning("Orange: bulk download failed: %s", e)
        return cache_path if cache_path.exists() else None


def _orange_build_delinquent_by_pin() -> dict[str, dict]:
    path = _orange_download()
    if not path:
        return {}

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().upper() if h else "" for h in next(rows)]
    col = {name: i for i, name in enumerate(header)}

    required = {"NAME1", "PARCEL #", "BASE TAX AMT"}
    if not required.issubset(col):
        logger.warning(
            "Orange: unexpected column layout (missing %s) — skipping",
            required - set(col),
        )
        return {}

    by_pin: dict[str, dict] = {}
    for row in rows:
        pin = str(row[col["PARCEL #"]] or "").strip()
        if not pin:
            continue
        try:
            amount = float(row[col["BASE TAX AMT"]] or 0)
        except (TypeError, ValueError):
            amount = 0.0
        by_pin[pin] = {
            "delinquent_amount": round(amount, 2),
            "delinquent_years": 1,  # annual snapshot has no per-year breakdown
            "owner": row[col["NAME1"]] or "",
        }
    logger.info("Orange: indexed %d delinquent parcels", len(by_pin))
    return by_pin


def _orange_lookup_pins_by_address(address: str) -> list[str]:
    """Return candidate PINs for an address — sometimes more than one

    (adjoining/resubdivided parcels can share a street address; confirmed
    live — e.g. "2823 Butler Rd" resolves to two distinct PINs, only one
    of which is actually delinquent), so the caller checks each against
    the delinquent index rather than assuming the first result is right.
    """
    key = _address_key(address)
    if not key:
        return []
    house_number, street_word = key
    where = f"ADDRESS1 LIKE '{house_number} {street_word}%'"
    params = {"where": where, "outFields": "PIN", "f": "json", "resultRecordCount": 5}
    try:
        resp = requests.get(
            ORANGE_PARCEL_URL, params=params, timeout=REQUEST_TIMEOUT, headers=_HEADERS
        )
        resp.raise_for_status()
        features = resp.json().get("features", [])
        return [f["attributes"].get("PIN") for f in features if f["attributes"].get("PIN")]
    except (requests.RequestException, ValueError) as e:
        logger.debug("Orange: parcel PIN lookup failed for '%s': %s", address, e)
        return []


def _apply_orange(notices: list[NoticeData]) -> None:
    by_pin = _orange_build_delinquent_by_pin()
    if not by_pin:
        logger.info("Orange: no delinquency data available this run")
        return
    enriched = 0
    for n in notices:
        time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))
        pin = next((p for p in _orange_lookup_pins_by_address(n.address) if p in by_pin), None)
        if not pin:
            continue
        result = by_pin[pin]
        n.tax_delinquent_amount = str(result["delinquent_amount"])
        n.tax_delinquent_years = str(result["delinquent_years"])
        if result.get("owner") and not n.tax_owner_name:
            n.tax_owner_name = result["owner"]
        n.parcel_id = n.parcel_id or pin
        enriched += 1
    logger.info("Orange: tax-delinquent %d/%d", enriched, len(notices))


# ── Mecklenburg County (unsupported) ─────────────────────────────────────

_mecklenburg_notified = False


def _mecklenburg_note_once() -> None:
    global _mecklenburg_notified
    if not _mecklenburg_notified:
        logger.info(
            "Mecklenburg: no automated tax delinquency source available "
            "(WAF-protected bill-search portal, no public API or bulk "
            "file) — skipped. See CLAUDE.md 'NC Tax Delinquency Enrichment' "
            "for the manual public-records-request option."
        )
        _mecklenburg_notified = True


# ── Entry Point ──────────────────────────────────────────────────────────


def _apply_index_lookup(notices: list[NoticeData], index: dict, label: str) -> None:
    if not index:
        logger.info("%s: no delinquency data available this run", label)
        return
    enriched = 0
    for n in notices:
        key = _address_key(n.address)
        if not key or key not in index:
            continue
        result = index[key]
        n.tax_delinquent_amount = str(result["delinquent_amount"])
        n.tax_delinquent_years = str(result["delinquent_years"])
        if result.get("owner") and not n.tax_owner_name:
            n.tax_owner_name = result["owner"]
        enriched += 1
    logger.info("%s: tax-delinquent %d/%d", label, enriched, len(notices))


def enrich_nc_tax_delinquency(notices: list[NoticeData]) -> None:
    """Enrich NC notices with county tax delinquency data.

    Dispatches per county — each NC county publishes delinquency data
    through a completely different system (see module docstring). Updates
    notices in-place with the same tax_delinquent_amount /
    tax_delinquent_years / tax_owner_name fields tax_enricher.py uses for
    Knox County, so downstream code (CSV export, DataSift mapping) needs
    no changes to pick up NC data.
    """
    by_county: dict[str, list[NoticeData]] = defaultdict(list)
    for n in notices:
        county = n.county.strip().lower()
        if county in NC_TAX_COUNTIES:
            by_county[county].append(n)

    if not by_county:
        return

    if "wake" in by_county:
        _apply_index_lookup(by_county["wake"], _wake_build_index(), "Wake")

    if "durham" in by_county:
        _apply_index_lookup(by_county["durham"], _durham_build_index(), "Durham")

    if "guilford" in by_county:
        _apply_guilford(by_county["guilford"])

    if "orange" in by_county:
        _apply_orange(by_county["orange"])

    if "mecklenburg" in by_county:
        _mecklenburg_note_once()
