"""Foreclosure timeline analysis report generator.

Builds a 5-sheet Excel workbook (Executive Summary, Timing Analysis, Zip Code
Analysis, Equity & Property, Property Data (Deduped)) analyzing how a
county's foreclosure notices move from publication to auction to sale —
matching the layout of the original Knox_Foreclosure_Analysis.xlsx report.

Two ways to get data in:

  --scrape --counties Wake Durham --months-back 12
      Runs a live historical foreclosure scrape of ncnotices.com (published
      NC foreclosure sale notices) via ncnotices_scraper.scrape_all_nc() +
      enrichment_pipeline.run_enrichment_pipeline() — the same call path
      python src/main.py nc-historical uses, scoped to notice_type=foreclosure.
      Covers the 5-county NC Triangle/Piedmont expansion market wired up in
      config.NC_SAVED_SEARCHES: Wake, Durham, Orange, Guilford, Mecklenburg.
      Each notice detail page is gated by a Cloudflare Turnstile solve, so
      this costs real time (and 2Captcha spend) — not instant, but lighter
      than the TN reCAPTCHA path since there's no login/session dance.

  --input output/some_export.csv [more.csv ...]
      Reuses CSV(s) already produced by any existing entry point (historical,
      nc-historical, ecourts-historical, daily, csv-import) via
      data_formatter.read_csv(). Works for any county/source already scraped
      — including TN Knox/Blount, or NC eCourts Special Proceedings filings
      once ecourts_scraper.py's Scrapfly bypass is confirmed working (see
      root CLAUDE.md's "NC eCourts Portal" section — not yet verified
      end-to-end). Since these exports are already deduplicated, the raw
      pre-dedup notice count and repeat-notice count aren't recoverable in
      this mode and are reported as "N/A".

Usage:
  python src/foreclosure_timeline_report.py --scrape --counties Wake Durham Orange Guilford Mecklenburg --months-back 12
  python src/foreclosure_timeline_report.py --input output/nc_notices_2026-07-20.csv --county-label Wake
"""

import argparse
import asyncio
import logging
import statistics
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parent))

from data_formatter import read_csv  # noqa: E402
from notice_parser import NoticeData  # noqa: E402
from property_registry import property_key  # noqa: E402

logger = logging.getLogger(__name__)

# ── Styling (matches the original Knox_Foreclosure_Analysis.xlsx look) ─────
NAVY = "1F4E79"
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=12, color="666666")
SECTION_FONT = Font(name="Calibri", bold=True, size=13, color=NAVY)
LABEL_FONT = Font(name="Calibri", bold=True, size=11)
VALUE_FONT = Font(name="Calibri", size=11)
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


# ── Small parsing helpers ───────────────────────────────────────────────────

def _f(val) -> float | None:
    """Parse a string/number field to float, tolerating blanks and $/, chars."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(",", "").replace("%", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _d(val) -> date | None:
    if not val:
        return None
    s = str(val).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _pct(n: int, total: int) -> str:
    return f"{(n / total * 100):.1f}%" if total else "0.0%"


def _money(v: float | None) -> str:
    return f"${v:,.0f}" if v is not None else ""


# ── Data loading ─────────────────────────────────────────────────────────

def load_from_csvs(paths: list[str], county_filter: list[str] | None) -> list[NoticeData]:
    notices: list[NoticeData] = []
    for pattern in paths:
        matched = sorted(Path().glob(pattern)) if any(ch in pattern for ch in "*?[") else [Path(pattern)]
        if not matched:
            logger.warning("No files matched: %s", pattern)
        for p in matched:
            notices.extend(read_csv(p))

    notices = [n for n in notices if n.notice_type == "foreclosure"]
    if county_filter:
        wanted = {c.lower() for c in county_filter}
        notices = [n for n in notices if n.county.lower() in wanted]
    return notices


def scrape_live(
    counties: list[str], months_back: int, max_notices: int, no_enrich: bool = False,
) -> tuple[list[NoticeData], list[NoticeData]]:
    """Run a live NC (ncnotices.com) historical foreclosure scrape + enrichment.

    Returns (raw_notices, final_notices) — raw is the pre-dedup list straight
    off the scraper (one row per notice detail page visited), final is post
    enrichment_pipeline (dedup + Zillow/Smarty enrichment; the TN-only
    enrichers — Knox Tax API, obituary/Ancestry deceased-owner research —
    are skipped since they don't apply to NC addresses, mirroring
    main.py's _run_nc_scrape_pipeline). Comparing the two lengths gives the
    "Total Raw Records" vs "Unique Properties" stats.

    no_enrich=True skips Zillow/Smarty/tax entirely — dedup and the
    vacant-land/entity-owner/commercial filters (all address/name-text based,
    no external calls) still run, since those are needed to get a clean
    property list. This trades the equity/value/MLS-status columns (blank in
    the output) for a fast raw notice-timeline snapshot with no API calls.
    """
    import config as cfg
    from enrichment_pipeline import PipelineOptions, run_enrichment_pipeline
    from ncnotices_scraper import scrape_all_nc

    searches = [s for s in cfg.NC_SAVED_SEARCHES if s.notice_type == "foreclosure"]
    if counties:
        wanted = {c.lower() for c in counties}
        searches = [s for s in searches if s.county.lower() in wanted]
    if not searches:
        raise SystemExit(
            f"No NC foreclosure saved search matches counties={counties}. "
            f"Available: {[s.county for s in cfg.NC_SAVED_SEARCHES if s.notice_type == 'foreclosure']}"
        )

    days_back = 30 * months_back
    logger.info("Scraping %d NC saved search(es), last %d days...", len(searches), days_back)

    raw_notices = asyncio.run(scrape_all_nc(
        mode="historical",
        searches=searches,
        days_back=days_back,
        max_notices=max_notices,
    ))
    logger.info("Scrape returned %d raw notices", len(raw_notices))

    opts = PipelineOptions(
        skip_parcel_lookup=True,
        skip_tax=True,
        skip_obituary=True,
        skip_ancestry=True,
        skip_smarty=no_enrich,
        skip_geocode=no_enrich,
        skip_zillow=no_enrich,
        skip_narrpr=no_enrich,
        source_label="foreclosure-timeline-report",
    )
    final_notices = run_enrichment_pipeline(list(raw_notices), opts)
    return raw_notices, final_notices


# ── Stat computation ────────────────────────────────────────────────────

BUCKET_ORDER = ["0-30 days", "31-60 days", "61-90 days", "91-120 days", "121-180 days", "180+ days"]
BUCKET_NOTES = {
    "0-30 days": "Fast movers — likely pre-foreclosure deals or auction purchases",
    "31-60 days": "Early outreach converts here — your mail/calls are hitting",
    "61-90 days": "Follow-up window — 2nd/3rd touch pays off",
    "91-120 days": "Post-auction cleanup — some cure, then sell anyway",
    "121-180 days": "Long-tail conversions — drip campaigns and patience",
    "180+ days": "These took 6+ months — persistence or REO sales",
}

VALUE_RANGES = [
    ("<$150K", lambda v: v < 150_000),
    ("$150K-$250K", lambda v: 150_000 <= v < 250_000),
    ("$250K-$400K", lambda v: 250_000 <= v < 400_000),
    ("$400K-$600K", lambda v: 400_000 <= v < 600_000),
    ("$600K+", lambda v: v >= 600_000),
]

OWNERSHIP_BUCKETS = [
    ("0-3 years", lambda y: y < 3,
     "Recent buyer — overleveraged or life event",
     "Empathy-first. \"We know this wasn't the plan.\""),
    ("3-5 years", lambda y: 3 <= y < 5,
     "Post-COVID buyer — rate/value squeeze",
     "Focus on getting them out whole. Avoid shame triggers."),
    ("5-10 years", lambda y: 5 <= y < 10,
     "Mid-term owner — divorce, medical, job loss",
     "Solution-oriented. \"We can close fast and give you options.\""),
    ("10-20 years", lambda y: 10 <= y < 20,
     "Long-term owner — aging, inherited debt",
     "Respect & trust. Slow play. Multiple touches."),
    ("20+ years", lambda y: y >= 20,
     "Legacy owner — major equity, emotional attachment",
     "Highest equity but hardest to move. Community proof works."),
]

STATUS_ACTIONS = {
    "sold": "These ALREADY sold — study for comps and sale price patterns",
    "off market": "HIGHEST PRIORITY — sitting on equity, no plan, no agent. Your #1 target.",
    "active": "Listed with agent — harder to get. Track for price drops and expireds.",
    "foreclosed": "Bank-owned now — approach as REO. Different negotiation.",
    "pending": "Under contract — watch for fall-throughs.",
    "for rent": "Investor-owned — may be open to selling the note or property.",
}
DEFAULT_ACTION = "Review individually — uncommon status, no default playbook."

STATUS_ORDER = ["Sold", "Off Market", "Active", "Foreclosed", "Pending", "For Rent"]


def outcome_bucket(status: str) -> str:
    """Bucket a raw mls_status value — blank is "Unknown" (no enrichment data),
    distinct from "Off Market" (a real Zillow-confirmed status). Conflating
    the two would make an unenriched or partially-enriched batch look like
    it's mostly off-market when really most records just have no data.
    """
    s = (status or "").strip().lower()
    if s in ("sold", "closed"):
        return "Sold"
    if s == "active":
        return "Active"
    if s in ("off market", "offmarket"):
        return "Off Market"
    if not s:
        return "Unknown"
    return "Other"


def bucket_days(days: float) -> str:
    if days <= 30:
        return "0-30 days"
    if days <= 60:
        return "31-60 days"
    if days <= 90:
        return "61-90 days"
    if days <= 120:
        return "91-120 days"
    if days <= 180:
        return "121-180 days"
    return "180+ days"


def compute_stats(
    notices: list[NoticeData],
    raw_count: int | None,
    repeat_count: int | None,
    county_label: str,
) -> dict:
    stats: dict = {"county_label": county_label}

    n = len(notices)
    stats["unique_properties"] = n
    stats["raw_count"] = raw_count if raw_count is not None else n
    stats["repeat_count"] = repeat_count
    # Zillow enrichment (mls_status/estimated_value) may have been skipped
    # entirely (--no-enrich, or a CSV that was never enriched) — in that case
    # outcome/equity/ownership-duration sections have nothing real to show
    # and should say so plainly instead of rendering misleading 0%/100% splits.
    stats["has_enrichment"] = any(
        (x.mls_status or "").strip() or (x.estimated_value or "").strip() for x in notices
    )

    # Date range
    notice_dates = [d for d in (_d(x.date_added) for x in notices) if d]
    if notice_dates:
        stats["date_range"] = f"{min(notice_dates):%b %Y} – {max(notice_dates):%b %Y}"
    else:
        stats["date_range"] = "N/A"

    # Notice-to-auction (days)
    n2a = []
    for x in notices:
        nd, ad = _d(x.date_added), _d(x.auction_date)
        if nd and ad and ad >= nd:
            n2a.append((ad - nd).days)
    stats["n2a"] = n2a
    if n2a:
        n2a_sorted = sorted(n2a)
        q = statistics.quantiles(n2a_sorted, n=4, method="inclusive") if len(n2a_sorted) >= 2 else [n2a_sorted[0]] * 3
        stats["n2a_min"] = min(n2a_sorted)
        stats["n2a_p25"] = q[0]
        stats["n2a_median"] = statistics.median(n2a_sorted)
        stats["n2a_p75"] = q[2]
        stats["n2a_max"] = max(n2a_sorted)
    else:
        stats["n2a_min"] = stats["n2a_p25"] = stats["n2a_median"] = stats["n2a_p75"] = stats["n2a_max"] = None

    # Outcome buckets
    outcome_counts = defaultdict(int)
    for x in notices:
        outcome_counts[outcome_bucket(x.mls_status)] += 1
    stats["outcome_counts"] = dict(outcome_counts)
    sold_n = outcome_counts.get("Sold", 0)
    off_market_n = outcome_counts.get("Off Market", 0)
    active_n = outcome_counts.get("Active", 0)
    unknown_n = outcome_counts.get("Unknown", 0)
    other_n = n - sold_n - off_market_n - active_n - unknown_n
    stats["sold_n"], stats["off_market_n"], stats["active_n"], stats["other_n"], stats["unknown_n"] = (
        sold_n, off_market_n, active_n, other_n, unknown_n,
    )

    # Notice-to-sale timing (sold properties whose recorded last-sold date
    # falls on/after the notice date — i.e. a resale that happened because of
    # this foreclosure, not the purchase that preceded it)
    n2s = []
    sold_before = sold_on = sold_after = 0
    for x in notices:
        if outcome_bucket(x.mls_status) != "Sold":
            continue
        nd, sd, ad = _d(x.date_added), _d(x.mls_last_sold_date), _d(x.auction_date)
        if nd and sd and sd >= nd:
            n2s.append((sd - nd).days)
            if ad:
                if sd < ad:
                    sold_before += 1
                elif sd == ad:
                    sold_on += 1
                else:
                    sold_after += 1
    stats["n2s"] = n2s
    stats["n2s_median"] = statistics.median(n2s) if n2s else None
    stats["sold_before"], stats["sold_on"], stats["sold_after"] = sold_before, sold_on, sold_after

    n2s_buckets = {b: 0 for b in BUCKET_ORDER}
    for days in n2s:
        n2s_buckets[bucket_days(days)] += 1
    stats["n2s_buckets"] = n2s_buckets

    # Marketing / peak windows (narrative, derived from the timing distribution)
    if stats["n2a_median"] is not None:
        stats["marketing_window_days"] = max(10, int(stats["n2a_median"] // 10) * 10)
    else:
        stats["marketing_window_days"] = None
    # Narrative "peak window" — the two buckets right after the initial
    # marketing window, where follow-up outreach tends to convert. Only
    # surfaced if there's actually sold-property data to back it up.
    stats["peak_window_label"] = "Days 30-90 post-notice" if n2s else None

    # Equity
    equities = [_f(x.equity_percent) for x in notices]
    equities = [e for e in equities if e is not None]
    stats["equity_all"] = equities
    stats["equity_median_all"] = statistics.median(equities) if equities else None
    stats["equity_high_count"] = sum(1 for e in equities if e >= 60)

    equity_by_outcome: dict[str, list[float]] = defaultdict(list)
    for x in notices:
        e = _f(x.equity_percent)
        if e is not None:
            equity_by_outcome[outcome_bucket(x.mls_status)].append(e)
    stats["equity_median_sold"] = (
        statistics.median(equity_by_outcome["Sold"]) if equity_by_outcome.get("Sold") else None
    )
    stats["equity_median_off_market"] = (
        statistics.median(equity_by_outcome["Off Market"]) if equity_by_outcome.get("Off Market") else None
    )

    # Ownership duration (years owned before the notice)
    years_owned = []
    for x in notices:
        nd, sd = _d(x.date_added), _d(x.mls_last_sold_date)
        if nd and sd and sd < nd:
            years_owned.append((nd - sd).days / 365.25)
    stats["years_owned"] = years_owned
    stats["avg_years_owned"] = statistics.mean(years_owned) if years_owned else None

    ownership_rows = []
    for label, predicate, profile, messaging in OWNERSHIP_BUCKETS:
        count = sum(1 for y in years_owned if predicate(y))
        ownership_rows.append((label, count, profile, messaging))
    stats["ownership_rows"] = ownership_rows

    # Zip code analysis
    by_zip: dict[str, list[NoticeData]] = defaultdict(list)
    for x in notices:
        z = (x.zip or "").strip()[:5]
        if z:
            by_zip[z].append(x)
    zip_rows = []
    for z, group in by_zip.items():
        g_sold = sum(1 for x in group if outcome_bucket(x.mls_status) == "Sold")
        g_active = sum(1 for x in group if outcome_bucket(x.mls_status) == "Active")
        g_equity = [e for e in (_f(x.equity_percent) for x in group) if e is not None]
        g_value = [v for v in (_f(x.estimated_value) for x in group) if v is not None]
        zip_rows.append({
            "zip": z,
            "total": len(group),
            "sold": g_sold,
            "active": g_active,
            "sold_pct": g_sold / len(group) if group else 0,
            "median_equity": statistics.median(g_equity) if g_equity else None,
            "median_value": statistics.median(g_value) if g_value else None,
        })
    zip_rows.sort(key=lambda r: -r["total"])
    # Rank-based tiers (top/middle/bottom third by volume) so this scales to
    # any county size instead of hardcoding Knox's own count thresholds.
    n_zips = len(zip_rows)
    for i, row in enumerate(zip_rows):
        if i < max(1, round(n_zips / 3)):
            row["tier"] = "TIER 1"
        elif i < max(2, round(2 * n_zips / 3)):
            row["tier"] = "TIER 2"
        else:
            row["tier"] = "TIER 3"
    stats["zip_rows"] = zip_rows

    # Value range buckets
    value_rows = []
    for label, predicate in VALUE_RANGES:
        group = [x for x in notices if _f(x.estimated_value) is not None and predicate(_f(x.estimated_value))]
        g_sold = sum(1 for x in group if outcome_bucket(x.mls_status) == "Sold")
        g_equity = [e for e in (_f(x.equity_percent) for x in group) if e is not None]
        value_rows.append({
            "label": label,
            "count": len(group),
            "sold": g_sold,
            "sold_pct": g_sold / len(group) if group else 0,
            "median_equity": statistics.median(g_equity) if g_equity else None,
        })
    stats["value_rows"] = value_rows
    if value_rows:
        biggest = max(value_rows, key=lambda r: r["count"])
        stats["biggest_value_range"] = biggest

    # Equity by MLS status (Equity & Property sheet)
    status_rows = []
    by_status: dict[str, list[NoticeData]] = defaultdict(list)
    for x in notices:
        label = (x.mls_status or "").strip() or "Unknown"
        by_status[label].append(x)
    ordered_labels = [s for s in STATUS_ORDER if s in by_status] + [
        s for s in by_status if s not in STATUS_ORDER
    ]
    for label in ordered_labels:
        group = by_status[label]
        g_equity = [e for e in (_f(x.equity_percent) for x in group) if e is not None]
        low = sum(1 for e in g_equity if e < 30)
        mid = sum(1 for e in g_equity if 30 <= e < 60)
        high = sum(1 for e in g_equity if e >= 60)
        action = STATUS_ACTIONS.get(label.lower(), DEFAULT_ACTION)
        status_rows.append({
            "label": label,
            "count": len(group),
            "median_equity": statistics.median(g_equity) if g_equity else None,
            "low": low, "mid": mid, "high": high,
            "action": action,
        })
    stats["status_rows"] = status_rows

    return stats


# ── Workbook building ───────────────────────────────────────────────────

def _title_row(ws: Worksheet, row: int, text: str, font: Font, span: int = 6) -> None:
    ws.cell(row=row, column=1, value=text).font = font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def build_workbook(stats: dict, notices: list[NoticeData], out_path: Path) -> None:
    wb = Workbook()
    _write_exec_summary(wb.active, stats)
    wb.active.title = "Executive Summary"
    _write_timing_analysis(wb.create_sheet("Timing Analysis"), stats)
    _write_zip_analysis(wb.create_sheet("Zip Code Analysis"), stats)
    _write_equity_property(wb.create_sheet("Equity & Property"), stats)
    _write_property_data(wb.create_sheet("Property Data (Deduped)"), notices)
    wb.save(out_path)


def _write_exec_summary(ws: Worksheet, s: dict) -> None:
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20

    row = 1
    _title_row(ws, row, f"{s['county_label']} Foreclosure Analysis", TITLE_FONT)
    row += 1
    _title_row(ws, row, "Marketing Intelligence Report", SUBTITLE_FONT)
    row += 2

    def kv(label: str, value) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=3, value=value).font = VALUE_FONT
        row += 1

    def section(title: str) -> None:
        nonlocal row
        _title_row(ws, row, title, SECTION_FONT, span=3)
        row += 1

    section("DATASET OVERVIEW")
    kv("Total Raw Records", s["raw_count"] if s["raw_count"] is not None else "N/A")
    kv("Unique Properties", s["unique_properties"])
    kv("Properties with Repeat Notices", s["repeat_count"] if s["repeat_count"] is not None else "N/A")
    kv("Date Range", s["date_range"])
    kv("County", s["county_label"])
    row += 1

    section("TIMING INTELLIGENCE")
    kv("Median Notice-to-Auction", f"{s['n2a_median']:.0f} days" if s["n2a_median"] is not None else "N/A")
    kv("Avg Notice-to-Auction", f"{statistics.mean(s['n2a']):.0f} days" if s["n2a"] else "N/A")
    kv("Median Notice-to-Sale (sold props)", f"{s['n2s_median']:.0f} days" if s["n2s_median"] is not None else "N/A")
    kv("Marketing Window", f"Day 1 – Day {s['marketing_window_days']}" if s["marketing_window_days"] else "N/A")
    kv("Peak Sale Window", s["peak_window_label"] or "N/A")
    row += 1

    section("OUTCOME BREAKDOWN")
    n = s["unique_properties"]
    if s["has_enrichment"]:
        kv("Sold", f"{s['sold_n']} ({_pct(s['sold_n'], n)})")
        kv("Off Market (unsold, not listed)", f"{s['off_market_n']} ({_pct(s['off_market_n'], n)})")
        kv("Active (currently listed)", f"{s['active_n']} ({_pct(s['active_n'], n)})")
        kv("Foreclosed / Pending / Other", f"{s['other_n']} ({_pct(s['other_n'], n)})")
        if s["unknown_n"]:
            kv("Unknown (not enriched)", f"{s['unknown_n']} ({_pct(s['unknown_n'], n)})")
    else:
        kv("(not available)", "Property enrichment (Zillow MLS status) was skipped for this run")
    row += 1

    section("EQUITY INSIGHTS")
    if s["has_enrichment"]:
        kv("Median Equity (all props)", f"{s['equity_median_all']:.1f}%" if s["equity_median_all"] is not None else "N/A")
        kv("Sold Props Median Equity", f"{s['equity_median_sold']:.1f}%" if s["equity_median_sold"] is not None else "N/A")
        kv("Off Market Median Equity", f"{s['equity_median_off_market']:.1f}%" if s["equity_median_off_market"] is not None else "N/A")
        kv("High Equity (60%+) Properties", f"{s['equity_high_count']} ({_pct(s['equity_high_count'], len(s['equity_all']))} of those with data)")
        kv("Avg Years Owned Before Foreclosure", f"{s['avg_years_owned']:.1f} years" if s["avg_years_owned"] is not None else "N/A")
    else:
        kv("(not available)", "Property enrichment (Zillow equity/value) was skipped for this run")
    row += 1

    section("KEY MARKETING TAKEAWAYS")
    for bullet in build_takeaways(s):
        ws.cell(row=row, column=1, value=bullet).alignment = WRAP_TOP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1


def build_takeaways(s: dict) -> list[str]:
    out = []
    i = 1
    if s["n2a_median"] is not None:
        out.append(
            f"{i}. YOUR WINDOW IS ~{s['n2a_median']:.0f} DAYS — From notice publication to auction is a "
            f"median of {s['n2a_median']:.0f} days. You need to make contact within the first 7-14 days to "
            "have any meaningful conversation before the sale."
        )
        i += 1
    n = s["unique_properties"]
    # Only claim a sold/not-sold split if most records actually have a known
    # status — a couple of enriched records out of a mostly-unknown batch
    # isn't enough to say "X% did not sell".
    if n and s["has_enrichment"] and s["unknown_n"] < n / 2:
        not_sold_pct = _pct(n - s["sold_n"], n)
        out.append(
            f"{i}. MOST PROPERTIES DON'T SELL — {not_sold_pct} of foreclosure-noticed properties did NOT "
            "sell on MLS. These are your targets: distressed owners sitting on equity with no plan."
        )
        i += 1
    if s["equity_median_sold"] is not None and s["equity_median_off_market"] is not None:
        out.append(
            f"{i}. SOLD PROPERTIES HAD LOW EQUITY — Median equity on sold properties was only "
            f"{s['equity_median_sold']:.1f}% vs {s['equity_median_off_market']:.1f}% for off-market. The "
            "high-equity owners are NOT listing — they need YOU to bring them a solution."
        )
        i += 1
    if s["repeat_count"] is not None and n:
        out.append(
            f"{i}. REPEAT NOTICES = URGENCY SIGNAL — {_pct(s['repeat_count'], n)} of properties had 2+ "
            "notices. Multiple notices mean the process is escalating. Prioritize properties on their 2nd "
            "or 3rd notice."
        )
        i += 1
    top_zips = s["zip_rows"][:5]
    if top_zips:
        total = sum(r["total"] for r in s["zip_rows"])
        top_share = _pct(sum(r["total"] for r in top_zips), total)
        zips_str = ", ".join(f"{r['zip']} ({r['total']} props)" for r in top_zips)
        out.append(
            f"{i}. TOP ZIP CODES TO TARGET — {zips_str}. These {len(top_zips)} zips account for "
            f"{top_share} of all foreclosure activity."
        )
        i += 1
    if s["peak_window_label"]:
        out.append(
            f"{i}. PEAK SALE WINDOW IS {s['peak_window_label'].upper()} — Of properties that did sell, "
            "this is the sweet spot before the auction forces their hand."
        )
        i += 1
    if s["avg_years_owned"] is not None:
        out.append(
            f"{i}. LONG-TERM OWNERS DOMINATE — Average ownership before foreclosure is "
            f"{s['avg_years_owned']:.1f} years. These aren't flippers in trouble — these are homeowners "
            "who've had a life event. Approach with empathy."
        )
        i += 1
    biggest = s.get("biggest_value_range")
    total_valued = sum(r["count"] for r in s["value_rows"]) if biggest else 0
    if biggest and total_valued:
        pct_share = _pct(biggest["count"], total_valued)
        out.append(
            f"{i}. PRICE SWEET SPOT IS {biggest['label']} — {pct_share} of properties fall in this "
            "range. Your comps and offers should be calibrated here."
        )
        i += 1
    return out


def _write_timing_analysis(ws: Worksheet, s: dict) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 55

    row = 1
    _title_row(ws, row, "FORECLOSURE TIMING BREAKDOWN", SECTION_FONT, span=5)
    row += 1
    _header_row(ws, row, ["Metric", "Value", "Marketing Implication", "", ""])
    row += 1

    def timing_row(metric, days, implication):
        nonlocal row
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=f"{days:.0f} days" if days is not None else "N/A")
        ws.cell(row=row, column=3, value=implication)
        row += 1

    timing_row("Min Notice-to-Auction", s["n2a_min"], "Some properties get almost no warning — speed matters")
    timing_row("25th Percentile", s["n2a_p25"], "A quarter of auctions happen within this window")
    timing_row("Median", s["n2a_median"], "Plan your marketing cycle around this per lead")
    timing_row("75th Percentile", s["n2a_p75"], "Most properties give you at least this long")
    timing_row("Max", s["n2a_max"], "Some run long — these often have postponements")
    row += 1

    _title_row(ws, row, "NOTICE-TO-SALE TIMING (Sold Properties Only)", SECTION_FONT, span=5)
    row += 1
    if not s["has_enrichment"]:
        ws.cell(row=row, column=1, value=(
            "Not available — requires Zillow MLS status/sale-date data, skipped for this run."
        )).font = NOTE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        return
    _header_row(ws, row, ["Time Bucket", "Properties Sold", "% of Sales", "Cumulative %", "What This Means"])
    row += 1
    total_sold = sum(s["n2s_buckets"].values())
    cumulative = 0
    for b in BUCKET_ORDER:
        count = s["n2s_buckets"][b]
        cumulative += count
        ws.cell(row=row, column=1, value=b)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=_pct(count, total_sold))
        ws.cell(row=row, column=4, value=_pct(cumulative, total_sold))
        ws.cell(row=row, column=5, value=BUCKET_NOTES[b])
        row += 1
    row += 1

    _title_row(ws, row, "SOLD RELATIVE TO AUCTION DATE", SECTION_FONT, span=5)
    row += 1
    _header_row(ws, row, ["Category", "Count", "% of Sold", "Implication", ""])
    row += 1
    total_relative = s["sold_before"] + s["sold_on"] + s["sold_after"]
    relative_rows = [
        ("Sold BEFORE auction", s["sold_before"],
         "Pre-foreclosure deals — your ideal exit. Owner sells to avoid auction."),
        ("Sold ON auction day", s["sold_on"],
         "Auction sale — bank or investor bought at courthouse steps."),
        ("Sold AFTER auction", s["sold_after"],
         "Post-auction sales — REO, short sale, or delayed close."),
    ]
    for label, count, implication in relative_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=_pct(count, total_relative))
        ws.cell(row=row, column=4, value=implication)
        row += 1


def _write_zip_analysis(ws: Worksheet, s: dict) -> None:
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["H"].width = 14

    row = 1
    _title_row(ws, row, "FORECLOSURE ACTIVITY BY ZIP CODE", SECTION_FONT, span=8)
    row += 2
    _header_row(ws, row, [
        "Zip Code", "Total Properties", "Sold", "Active", "Sold %",
        "Median Equity %", "Median Est. Value", "Priority Tier",
    ])
    row += 1
    tier_min: dict[str, int] = {}
    tier_max: dict[str, int] = {}
    for r in s["zip_rows"]:
        ws.cell(row=row, column=1, value=r["zip"])
        ws.cell(row=row, column=2, value=r["total"])
        ws.cell(row=row, column=3, value=r["sold"])
        ws.cell(row=row, column=4, value=r["active"])
        ws.cell(row=row, column=5, value=f"{r['sold_pct']*100:.1f}%")
        ws.cell(row=row, column=6, value=f"{r['median_equity']:.1f}%" if r["median_equity"] is not None else "—")
        ws.cell(row=row, column=7, value=_money(r["median_value"]) or "—")
        ws.cell(row=row, column=8, value=r["tier"])
        tier_min[r["tier"]] = min(tier_min.get(r["tier"], r["total"]), r["total"])
        tier_max[r["tier"]] = max(tier_max.get(r["tier"], r["total"]), r["total"])
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="TIER DEFINITIONS:").font = LABEL_FONT
    row += 1
    parts = []
    for tier in ("TIER 1", "TIER 2", "TIER 3"):
        if tier in tier_min:
            lo, hi = tier_min[tier], tier_max[tier]
            span = f"{lo}+" if tier == "TIER 1" else f"{lo}-{hi}"
            role = {
                "TIER 1": "highest volume — mail/call first",
                "TIER 2": "secondary targets",
                "TIER 3": "cherry-pick high-equity deals",
            }[tier]
            parts.append(f"{tier} = {span} properties ({role})")
    ws.cell(row=row, column=1, value=". ".join(parts) + ".").alignment = WRAP_TOP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)


def _write_equity_property(ws: Worksheet, s: dict) -> None:
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 55

    row = 1
    _title_row(ws, row, "EQUITY ANALYSIS BY OUTCOME", SECTION_FONT, span=7)
    row += 1
    if not s["has_enrichment"]:
        ws.cell(row=row, column=1, value=(
            "Property enrichment (Zillow MLS status/equity/value) was skipped for this run — "
            "this sheet only has data when --scrape runs without --no-enrich."
        )).font = NOTE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        return
    _header_row(ws, row, [
        "MLS Status", "Count", "Median Equity %", "<30% Equity",
        "30-60% Equity", "60%+ Equity", "Marketing Action",
    ])
    row += 1
    for r in s["status_rows"]:
        ws.cell(row=row, column=1, value=r["label"])
        ws.cell(row=row, column=2, value=r["count"])
        ws.cell(row=row, column=3, value=f"{r['median_equity']:.1f}%" if r["median_equity"] is not None else "—")
        ws.cell(row=row, column=4, value=r["low"] if r["median_equity"] is not None else "—")
        ws.cell(row=row, column=5, value=r["mid"] if r["median_equity"] is not None else "—")
        ws.cell(row=row, column=6, value=r["high"] if r["median_equity"] is not None else "—")
        ws.cell(row=row, column=7, value=r["action"])
        row += 1
    row += 1

    _title_row(ws, row, "OWNERSHIP DURATION BEFORE FORECLOSURE", SECTION_FONT, span=7)
    row += 1
    _header_row(ws, row, ["Years Owned", "Count", "% of Total", "Owner Profile", "Messaging Approach", "", ""])
    row += 1
    total_owned = sum(c for _, c, _, _ in s["ownership_rows"])
    for label, count, profile, messaging in s["ownership_rows"]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=_pct(count, total_owned))
        ws.cell(row=row, column=4, value=profile)
        ws.cell(row=row, column=5, value=messaging)
        row += 1
    row += 1

    _title_row(ws, row, "PROPERTY VALUE RANGES", SECTION_FONT, span=7)
    row += 1
    _header_row(ws, row, ["Value Range", "Count", "Sold", "Sold %", "Median Equity", "", ""])
    row += 1
    for r in s["value_rows"]:
        ws.cell(row=row, column=1, value=r["label"])
        ws.cell(row=row, column=2, value=r["count"])
        ws.cell(row=row, column=3, value=r["sold"])
        ws.cell(row=row, column=4, value=_pct(r["sold"], r["count"]))
        ws.cell(row=row, column=5, value=f"{r['median_equity']:.1f}%" if r["median_equity"] is not None else "—")
        row += 1


PROPERTY_COLUMNS = [
    ("Owner Name", "owner_name"),
    ("Address", "address"),
    ("City", "city"),
    ("Zip", "zip"),
    ("First Notice", "date_added"),
    ("Auction Date", "auction_date"),
    ("MLS Status", "mls_status"),
    ("List Price", "mls_listing_price"),
    ("Last Sold Date", "mls_last_sold_date"),
    ("Last Sold Price", "mls_last_sold_price"),
    ("Est. Value", "estimated_value"),
    ("Est. Equity", "estimated_equity"),
    ("Equity %", "equity_percent"),
    ("Type", "property_type"),
    ("Beds", "bedrooms"),
    ("Baths", "bathrooms"),
    ("SqFt", "sqft"),
    ("Year Built", "year_built"),
    ("Vacant", "vacant"),
]


def _write_property_data(ws: Worksheet, notices: list[NoticeData]) -> None:
    _header_row(ws, 1, [label for label, _ in PROPERTY_COLUMNS])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    for i, notice in enumerate(notices, start=2):
        for c, (_, field) in enumerate(PROPERTY_COLUMNS, start=1):
            ws.cell(row=i, column=c, value=getattr(notice, field, ""))


# ── Repeat-notice detection (raw, pre-dedup) ────────────────────────────

def count_repeats(raw_notices: list[NoticeData]) -> int:
    """Count properties that appeared under 2+ raw notices before dedup."""
    counts: dict[str, int] = defaultdict(int)
    for n in raw_notices:
        if n.address.strip():
            counts[property_key(n)] += 1
    return sum(1 for c in counts.values() if c >= 2)


# ── CLI ──────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--scrape", action="store_true", help="Run a live NC (ncnotices.com) historical foreclosure scrape")
    parser.add_argument("--counties", nargs="+", help="Counties to scrape (NC: Wake, Durham, Orange, Guilford, Mecklenburg)")
    parser.add_argument("--months-back", type=int, default=12, help="Lookback window for --scrape (default: 12)")
    parser.add_argument("--max-notices", type=int, default=0, help="Cap notices scraped (0 = no cap)")
    parser.add_argument(
        "--no-enrich", action="store_true",
        help="Skip Zillow/Smarty/NARRPR entirely — fast raw notice-timeline snapshot, "
             "no property value/equity/MLS-status data (only applies with --scrape)",
    )
    parser.add_argument("--input", nargs="+", help="Existing CSV(s) to load instead of scraping")
    parser.add_argument("--county-label", default=None, help="County label for the report title (input mode)")
    parser.add_argument("--output", default=None, help="Output .xlsx path")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    if args.scrape:
        if not args.counties:
            parser.error("--scrape requires --counties")
        raw_notices, final_notices = scrape_live(
            args.counties, args.months_back, args.max_notices, no_enrich=args.no_enrich,
        )
        final_notices = [n for n in final_notices if n.notice_type == "foreclosure"]
        raw_count = len(raw_notices)
        repeat_count = count_repeats(raw_notices)
        county_label = " & ".join(args.counties)
    elif args.input:
        final_notices = load_from_csvs(args.input, args.counties)
        raw_count = None
        repeat_count = None
        county_label = args.county_label or (" & ".join(args.counties) if args.counties else "County")
    else:
        parser.error("Provide either --scrape --counties ... or --input <csv...>")
        return

    if not final_notices:
        logger.error("No foreclosure notices to analyze — nothing written")
        sys.exit(1)

    stats = compute_stats(final_notices, raw_count, repeat_count, county_label)

    if args.output:
        out_path = Path(args.output)
    else:
        safe_label = county_label.replace(" ", "_").replace("&", "and")
        out_path = Path("output") / f"{safe_label}_Foreclosure_Analysis_{datetime.now():%Y%m%d}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    build_workbook(stats, final_notices, out_path)
    logger.info("Report written: %s", out_path)


if __name__ == "__main__":
    main()
